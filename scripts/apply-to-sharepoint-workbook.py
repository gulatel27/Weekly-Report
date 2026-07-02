import argparse
import copy
import datetime as dt
import shutil
from pathlib import Path
import sys

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


DEFAULT_SHEET_NAME = "DS-2팀"
OUTPUTS_DIR = Path("outputs")


def log(message: str) -> None:
    print(f"[{dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {message}")


def find_latest_report() -> Path:
    candidates = [
        path
        for path in OUTPUTS_DIR.glob("락플레이스-DS2_주간보고_*.xlsx")
        if not path.name.startswith("~$") and "빈폼" not in path.name
    ]
    candidates.sort(key=lambda path: path.stat().st_mtime, reverse=True)
    if not candidates:
        raise FileNotFoundError(f"최종 보고서 파일을 찾지 못했습니다: {OUTPUTS_DIR}")
    return candidates[0]


def get_source_sheet(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    if len(wb.worksheets) == 1:
        return wb.worksheets[0]
    raise KeyError(f"소스 파일에서 시트를 찾지 못했습니다: {sheet_name}")


def create_replacement_sheet(target_wb, sheet_name: str):
    if sheet_name in target_wb.sheetnames:
        old_sheet = target_wb[sheet_name]
        sheet_index = target_wb.worksheets.index(old_sheet)
        sheet_state = old_sheet.sheet_state
        target_wb.remove(old_sheet)
        new_sheet = target_wb.create_sheet(sheet_name, sheet_index)
        new_sheet.sheet_state = sheet_state
        return new_sheet
    return target_wb.create_sheet(sheet_name)


def copy_dimensions(source_ws, target_ws) -> None:
    target_ws.sheet_format = copy.copy(source_ws.sheet_format)

    for key, dimension in source_ws.column_dimensions.items():
        target_ws.column_dimensions[key] = copy.copy(dimension)
        target_ws.column_dimensions[key].worksheet = target_ws

    for key, dimension in source_ws.row_dimensions.items():
        target_ws.row_dimensions[key] = copy.copy(dimension)
        target_ws.row_dimensions[key].worksheet = target_ws


def copy_cell(source_cell, target_cell) -> None:
    if isinstance(source_cell, MergedCell):
        return

    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell._style = copy.copy(source_cell._style)
    target_cell.font = copy.copy(source_cell.font)
    target_cell.fill = copy.copy(source_cell.fill)
    target_cell.border = copy.copy(source_cell.border)
    target_cell.alignment = copy.copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy.copy(source_cell.protection)

    if source_cell.hyperlink:
        target_cell._hyperlink = copy.copy(source_cell.hyperlink)
    if source_cell.comment:
        target_cell.comment = copy.copy(source_cell.comment)


def copy_sheet(source_ws, target_ws) -> None:
    target_ws.title = source_ws.title
    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    target_ws.sheet_view.zoomScale = source_ws.sheet_view.zoomScale or 100
    target_ws.sheet_view.zoomScaleNormal = source_ws.sheet_view.zoomScaleNormal or 100
    target_ws.sheet_view.zoomScalePageLayoutView = source_ws.sheet_view.zoomScalePageLayoutView or 100
    target_ws.sheet_properties = copy.copy(source_ws.sheet_properties)
    target_ws.page_margins = copy.copy(source_ws.page_margins)
    target_ws.page_setup = copy.copy(source_ws.page_setup)
    target_ws.print_options = copy.copy(source_ws.print_options)

    copy_dimensions(source_ws, target_ws)

    for row in range(1, source_ws.max_row + 1):
        for col in range(1, source_ws.max_column + 1):
            copy_cell(source_ws.cell(row, col), target_ws.cell(row, col))

    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    if source_ws.auto_filter.ref:
        target_ws.auto_filter.ref = source_ws.auto_filter.ref
    target_ws.sheet_properties.tabColor = source_ws.sheet_properties.tabColor


def build_output_path(target_path: Path) -> Path:
    return target_path.with_name(f"{target_path.stem}_DS2업데이트{target_path.suffix}")


def apply_report_to_target(source_path: Path, target_path: Path, output_path: Path, sheet_name: str) -> Path:
    source_wb = load_workbook(source_path)
    source_ws = get_source_sheet(source_wb, sheet_name)

    target_wb = load_workbook(target_path)
    target_ws = create_replacement_sheet(target_wb, sheet_name)
    copy_sheet(source_ws, target_ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    target_wb.save(output_path)
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="생성된 DS-2팀 주간보고 시트를 SharePoint 주간보고 엑셀 파일에 직접 반영"
    )
    parser.add_argument("--target", type=Path, required=True, help="SharePoint 원본 xlsx 파일 경로")
    parser.add_argument("--source", type=Path, default=None, help="생성된 DS-2팀 최종 보고서 xlsx")
    parser.add_argument("--output", type=Path, default=None, help="저장할 결과 xlsx 경로")
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME, help="교체할 시트명")
    parser.add_argument(
        "--in-place",
        action="store_true",
        help="대상 파일을 직접 덮어씁니다. 덮어쓰기 전 같은 폴더에 백업을 만듭니다.",
    )
    return parser.parse_args()


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    args = parse_args()
    source_path = args.source or find_latest_report()
    target_path = args.target
    if not source_path.exists():
        raise FileNotFoundError(f"소스 파일을 찾지 못했습니다: {source_path}")
    if not target_path.exists():
        raise FileNotFoundError(f"대상 파일을 찾지 못했습니다: {target_path}")

    if args.in_place:
        timestamp = dt.datetime.now().strftime("%Y%m%d%H%M%S")
        backup_path = target_path.with_name(f"{target_path.stem}_backup_{timestamp}{target_path.suffix}")
        shutil.copy2(target_path, backup_path)
        output_path = target_path
        log(f"백업 생성: {backup_path}")
    else:
        output_path = args.output or build_output_path(target_path)
        if output_path.resolve() == target_path.resolve():
            raise ValueError("--in-place 없이 대상 파일과 같은 경로에 저장할 수 없습니다.")

    log(f"소스 보고서: {source_path}")
    log(f"대상 파일: {target_path}")
    log(f"저장 파일: {output_path}")
    result_path = apply_report_to_target(source_path, target_path, output_path, args.sheet)
    log(f"SharePoint 업로드용 파일 생성 완료: {result_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
