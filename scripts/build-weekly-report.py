import argparse
import copy
import datetime as dt
import json
import os
from pathlib import Path
import re
import sys
import urllib.error
import urllib.request

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Border, Side


KEEP_SHEETS = ["DS-2팀"]
TEAM_NAME = "DS 2T"
TEMPLATE_GLOB = "락플레이스-지원부문_주간보고_*.xlsx"
DOWNLOADS_DIR = Path("D:/Downloads")
DEFAULT_LLM_CONFIG = Path("config/llm.local.json")
TIME_LABELS = ["주간", "야간", "심야", "휴일", "총합계"]
KIND_LABELS = ["설치", "지원", "점검", "Presales", "파견"]
LLM_USAGE_TOTALS = {"requests": 0, "prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}
ISSUE_ROW_WRAP_UNITS = 190
ISSUE_ROW_BASE_HEIGHT = 18
ISSUE_ROW_MAX_HEIGHT = 240
ISSUE_HIGHLIGHT_COLOR = "FF0000"


def log(message: str) -> None:
    print(f"[{dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {message}")


def this_week_range(today: dt.date | None = None) -> tuple[dt.date, dt.date]:
    today = today or dt.date.today()
    monday = today - dt.timedelta(days=today.weekday())
    sunday = monday + dt.timedelta(days=6)
    return monday, sunday


def parse_date_arg(value: str | None) -> dt.date | None:
    if not value:
        return None
    text = value.strip()
    for fmt in ("%Y-%m-%d", "%Y%m%d"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"날짜 형식은 YYYY-MM-DD 또는 YYYYMMDD 이어야 합니다: {value}")


def load_llm_config(config_path: Path | None = None) -> dict | None:
    summary_mode = str(os.environ.get("REPORT_SUMMARY_MODE") or "").strip().lower()
    if summary_mode in {"local", "2"}:
        log("요약 방식 선택: 로컬 요약을 사용합니다. LLM API는 호출하지 않습니다.")
        return None
    if summary_mode and summary_mode not in {"llm", "1"}:
        log(f"알 수 없는 요약 방식이라 로컬 요약을 사용합니다: {summary_mode}")
        return None

    raw_path = config_path or Path(os.environ.get("LLM_CONFIG_PATH") or DEFAULT_LLM_CONFIG)
    if not raw_path.exists():
        if summary_mode in {"llm", "1"}:
            log(f"LLM 요약을 선택했지만 설정 파일이 없어 로컬 요약을 사용합니다: {raw_path}")
        return None
    with raw_path.open("r", encoding="utf-8-sig") as file:
        config = json.load(file)
    if not config.get("enabled", False):
        return None
    api_key = str(config.get("api_key") or os.environ.get("OPENAI_API_KEY") or "").strip()
    if not api_key:
        log(f"LLM 설정 파일은 있으나 api_key가 비어 있어 로컬 요약을 사용합니다: {raw_path}")
        return None
    provider = str(config.get("provider") or "").strip().lower()
    if provider != "openai":
        log(f"지원하지 않는 LLM provider라 로컬 요약을 사용합니다: {provider}")
        return None
    config["api_key"] = api_key
    config["provider"] = provider
    config["base_url"] = str(config.get("base_url") or "https://api.openai.com/v1").rstrip("/")
    config["model"] = str(config.get("model") or "gpt-4o-mini")
    config["timeout_seconds"] = int(config.get("timeout_seconds") or 30)
    config["temperature"] = float(config.get("temperature", 0.2))
    config["max_tokens"] = int(config.get("max_tokens") or 500)
    return config


def find_latest_template() -> Path:
    candidates = sorted(
        DOWNLOADS_DIR.glob(TEMPLATE_GLOB),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(f"템플릿 파일을 찾지 못했습니다: {DOWNLOADS_DIR / TEMPLATE_GLOB}")
    return candidates[0]


def find_latest_download() -> Path:
    candidates = sorted(
        DOWNLOADS_DIR.glob("주간업무보고*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(f"다운로드 파일을 찾지 못했습니다: {DOWNLOADS_DIR / '주간업무보고*.xlsx'}")
    return candidates[0]


def unique_output_path(path: Path) -> Path:
    timestamp = dt.datetime.now().strftime("%Y%m%d%H%M%S")
    for index in range(1, 100):
        suffix = f"_{timestamp}" if index == 1 else f"_{timestamp}_{index}"
        candidate = path.with_name(f"{path.stem}{suffix}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"사용 가능한 출력 파일명을 만들지 못했습니다: {path}")


def output_path_is_writable(path: Path) -> bool:
    if not path.exists():
        return True
    try:
        with path.open("r+b"):
            return True
    except PermissionError:
        return False


def prepare_output_path(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    if output_path_is_writable(path):
        return path
    fallback_path = unique_output_path(path)
    log(f"저장 대상 파일이 열려 있어 새 파일명으로 저장합니다: {fallback_path}")
    return fallback_path


def save_workbook_with_fallback(wb, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(output_path)
        return output_path
    except PermissionError:
        fallback_path = unique_output_path(output_path)
        log(f"저장 대상 파일에 접근할 수 없어 새 파일명으로 저장합니다: {fallback_path}")
        wb.save(fallback_path)
        return fallback_path


def clear_cell(cell) -> None:
    if isinstance(cell, MergedCell):
        return
    cell.value = None


def clear_row(ws, row_number: int) -> None:
    for col in range(1, ws.max_column + 1):
        clear_cell(ws.cell(row_number, col))


def clear_whitespace_values(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.strip() == "":
                cell.value = None


def find_section_rows(ws) -> dict[int, int]:
    sections: dict[int, int] = {}
    patterns = {
        1: re.compile(r"^\s*1[\.,](?=\s|$)"),
        2: re.compile(r"^\s*2[\.,](?=\s|$)"),
        3: re.compile(r"^\s*3[\.,](?=\s|$)"),
        4: re.compile(r"^\s*4[\.,](?=\s|$)"),
        5: re.compile(r"^\s*5[\.,](?=\s|$)"),
        6: re.compile(r"^\s*6[\.,](?=\s|$)"),
    }
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if not isinstance(value, str):
            continue
        text = value.strip()
        for section_no, pattern in patterns.items():
            if section_no not in sections and pattern.search(text):
                sections[section_no] = row
                break
    return sections


def next_section_start(sections: dict[int, int], section_no: int, fallback: int) -> int:
    later = [row for no, row in sections.items() if no > section_no]
    return min(later) if later else fallback + 1


def row_values(ws, row_number: int) -> list:
    return [ws.cell(row_number, col).value for col in range(1, ws.max_column + 1)]


def is_summary_label_row(values: list) -> bool:
    non_empty = [value for value in values if value not in (None, "")]
    if not non_empty:
        return True
    return all(isinstance(value, str) and not value.startswith("=") for value in non_empty)


def is_section1_structure_row(value) -> bool:
    if not isinstance(value, str):
        return False
    text = value.strip()
    return (
        text.startswith("(1)")
        or text.startswith("(2)")
        or text.startswith("- 내부 기술 세미나")
        or text.startswith("- 자격증")
    )


def find_table_header_row(ws, start_row: int, end_row: int) -> int | None:
    for row in range(start_row + 1, end_row):
        values = row_values(ws, row)
        labels = {str(value).strip() for value in values if value not in (None, "")}
        if "일시" in labels and ("담당" in labels or "내용" in labels or "엔드유저" in labels):
            return row
    return None


def blank_report_sheet(ws) -> None:
    sections = find_section_rows(ws)
    missing = [section_no for section_no in range(1, 7) if section_no not in sections]
    if missing:
        raise ValueError(f"{ws.title} 시트에서 섹션을 찾지 못했습니다: {missing}")

    start = sections[1] + 1
    end = next_section_start(sections, 1, ws.max_row)
    for row in range(start, end):
        if not is_section1_structure_row(ws.cell(row, 1).value):
            clear_row(ws, row)

    start = sections[2] + 1
    end = next_section_start(sections, 2, ws.max_row)
    for row in range(start, end):
        clear_row(ws, row)

    start = sections[3] + 1
    end = next_section_start(sections, 3, ws.max_row)
    for row in range(start, end):
        if not is_summary_label_row(row_values(ws, row)):
            clear_row(ws, row)

    for section_no in (4, 5):
        start = sections[section_no] + 1
        end = next_section_start(sections, section_no, ws.max_row)
        header_row = find_table_header_row(ws, sections[section_no], end)
        for row in range(start, end):
            if row != header_row:
                clear_row(ws, row)

    start = sections[6] + 1
    for row in range(start, ws.max_row + 1):
        clear_row(ws, row)


def build_blank_report(template_path: Path):
    wb = load_workbook(template_path)
    for sheet_name in list(wb.sheetnames):
        if sheet_name not in KEEP_SHEETS:
            del wb[sheet_name]
    ws = wb["DS-2팀"]
    ws["A4"] = '= "보고일자 " & TEXT(NOW(), "yyyy-mm-dd")'
    set_sheet_zoom(ws)
    blank_report_sheet(ws)
    normalize_section_title_numbers(ws)
    clear_whitespace_values(ws)
    return wb


def set_sheet_zoom(ws, zoom: int = 100) -> None:
    ws.sheet_view.zoomScale = zoom
    ws.sheet_view.zoomScaleNormal = zoom
    ws.sheet_view.zoomScalePageLayoutView = zoom


def normalize_section_title_numbers(ws) -> None:
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row, 1)
        if not isinstance(cell.value, str):
            continue
        cell.value = re.sub(r"^(\s*[23])\s*,", r"\1.", cell.value, count=1)


def apply_middle_vertical_alignment(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            alignment = copy.copy(cell.alignment)
            alignment.vertical = "center"
            cell.alignment = alignment


def normalize_header(value) -> str:
    return re.sub(r"\s+", "", str(value or "").strip())


def find_header_row(ws, required_headers: set[str]) -> tuple[int, dict[str, int]]:
    normalized_required = {normalize_header(header) for header in required_headers}
    for row in range(1, min(ws.max_row, 20) + 1):
        headers: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            key = normalize_header(ws.cell(row, col).value)
            if key:
                headers[key] = col
        if normalized_required.issubset(headers):
            return row, headers
    raise ValueError(f"{ws.title} 시트에서 헤더를 찾지 못했습니다: {required_headers}")


def filter_downloaded_workbook(source_path: Path, team_name: str = TEAM_NAME) -> None:
    wb = load_workbook(source_path)

    weekly = wb["주간업무보고"]
    header_row, headers = find_header_row(weekly, {"부서명"})
    dept_col = headers["부서명"]
    for row in range(weekly.max_row, header_row, -1):
        if str(weekly.cell(row, dept_col).value or "").strip() != team_name:
            weekly.delete_rows(row, 1)

    if "근무-보상시간 통계" in wb.sheetnames:
        stats = wb["근무-보상시간 통계"]
        stats_header_row, stats_headers = find_header_row(stats, {"팀명"})
        team_col = stats_headers["팀명"]
        for row in range(stats.max_row, stats_header_row, -1):
            if str(stats.cell(row, team_col).value or "").strip() != team_name:
                stats.delete_rows(row, 1)

    wb.save(source_path)


def parse_datetime(value) -> dt.datetime | None:
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time())
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y.%m.%d %H:%M"):
        try:
            return dt.datetime.strptime(text, fmt)
        except ValueError:
            pass
    match = re.search(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})(?:\s+(\d{1,2}):(\d{2}))?", text)
    if not match:
        return None
    hour = int(match.group(4) or 0)
    minute = int(match.group(5) or 0)
    return dt.datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)), hour, minute)


def extract_kind(text: str) -> str | None:
    match = re.search(r'종류\s*:\s*["“]?([^",\r\n]+)["”]?', text or "", re.IGNORECASE)
    if not match:
        return None
    raw = match.group(1).strip()
    for label in KIND_LABELS:
        if raw.lower() == label.lower():
            return label
    return None


def extract_support_content(text: str) -> str:
    if not text:
        return ""
    match = re.search(r'지원내용\s*:\s*"(.*?)"\s*,', text, re.DOTALL)
    if match:
        return re.sub(r"\s+", " ", match.group(1)).strip()
    lines = [line.strip() for line in str(text).splitlines() if line.strip()]
    return lines[0] if lines else ""


def has_issue_yes(text: str) -> bool:
    return bool(re.search(r'이슈\s*:\s*["“]?\s*yes\s*["”]?', text or "", re.IGNORECASE))


def extract_detail_content(text: str) -> str:
    if not text:
        return ""
    match = re.search(r'상세내용\s*:\s*"(.*?)"\s*,\s*이슈\s*:', text, re.DOTALL | re.IGNORECASE)
    if match:
        return match.group(1).strip()
    match = re.search(r'상세내용\s*:\s*"(.*?)"', text, re.DOTALL | re.IGNORECASE)
    return match.group(1).strip() if match else ""


def summarize_detail(detail: str, fallback: str = "", max_lines: int = 4) -> list[str]:
    source = detail or fallback or ""
    cleaned_lines = []
    seen = set()
    skip_exact = {
        "구성",
        "구성:",
        "지원내용",
        "지원내용:",
        "이슈사항 or 요청사항",
        "이슈사항 or 요청사항:",
        "요청사항",
        "요청사항:",
        "원인",
        "원인:",
        "조치사항",
        "조치사항:",
        "기타 특이사항",
        "기타 특이사항:",
    }
    for raw_line in str(source).splitlines():
        line = raw_line.strip()
        line = re.sub(r"^[#\-\*\s]+", "", line)
        line = re.sub(r"\s+", " ", line).strip()
        normalized_label = re.sub(r"\s+", "", line).rstrip(":：")
        if not line or line in skip_exact or line.upper() in {"N/A", "NA", "NO"}:
            continue
        if normalized_label in {re.sub(r"\s+", "", item).rstrip(":：") for item in skip_exact}:
            continue
        if re.fullmatch(r"[가-힣A-Za-z0-9 /_-]{1,30}\s*[:：]", line):
            continue
        if len(line) > 115:
            line = line[:112].rstrip() + "..."
        key = line.lower()
        if key in seen:
            continue
        seen.add(key)
        cleaned_lines.append(line)
        if len(cleaned_lines) >= max_lines:
            break
    return cleaned_lines or ([fallback] if fallback else [])


def normalize_issue_key(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def clean_issue_text(raw_line: str) -> str:
    line = raw_line.strip()
    line = re.sub(r"^[#\-\*\s]+", "", line)
    line = re.sub(r"\s+", " ", line).strip()
    return line


def is_meaningful_issue_text(line: str) -> bool:
    normalized = re.sub(r"\s+", "", line).strip().upper()
    if not normalized or normalized in {"N/A", "NA", "NO", "없음", "해당없음"}:
        return False
    return not re.fullmatch(r"[가-힣A-Za-z0-9 /_-]{1,30}\s*[:：]", line)


def detail_label(raw_label: str) -> str | None:
    normalized = re.sub(r"\s+", "", raw_label).lower()
    aliases = {
        "지원내용": "지원 내용",
        "구성": "구성",
        "이슈사항or요청사항": "요청/이슈",
        "이슈사항": "요청/이슈",
        "요청사항": "요청/이슈",
        "원인": "원인",
        "조치사항": "조치",
        "조치": "조치",
        "기타특이사항": "기타",
        "특이사항": "기타",
    }
    return aliases.get(normalized)


def parse_detail_sections(detail: str) -> dict[str, list[str]]:
    sections: dict[str, list[str]] = {"개요": []}
    current = "개요"
    for raw_line in str(detail or "").splitlines():
        line = clean_issue_text(raw_line)
        if not line:
            continue
        match = re.match(r"^(.{1,30}?)[\s]*[:：]\s*(.*)$", line)
        if match:
            label = detail_label(match.group(1))
            if label:
                current = label
                sections.setdefault(current, [])
                rest = clean_issue_text(match.group(2))
                if is_meaningful_issue_text(rest):
                    sections[current].append(rest)
                continue
        if is_meaningful_issue_text(line):
            sections.setdefault(current, []).append(line)

    for label, lines in list(sections.items()):
        deduped = []
        seen = set()
        for line in lines:
            key = normalize_issue_key(line)
            if key in seen:
                continue
            seen.add(key)
            deduped.append(line)
        sections[label] = deduped
    return sections


def shorten_issue_text(value: str, max_chars: int | None = None) -> str:
    text = clean_issue_text(value)
    if max_chars is not None and len(text) > max_chars:
        text = text[: max_chars - 3].rstrip() + "..."
    return text


def compact_issue_values(values: list[str], max_items: int = 2, max_chars: int | None = None) -> str:
    selected = []
    for value in values:
        text = shorten_issue_text(value, max_chars=max_chars)
        if not is_meaningful_issue_text(text):
            continue
        selected.append(text)
        if len(selected) >= max_items:
            break
    if not selected:
        return ""

    compacted = " / ".join(selected)
    if max_chars is not None and len(compacted) > max_chars:
        compacted = compacted[: max_chars - 3].rstrip() + "..."
    return compacted


def issue_detail_lines(record: dict, max_lines: int = 5) -> list[str]:
    support = extract_support_content(record["지원내역"])
    sections = parse_detail_sections(record["상세내용"])
    lines = []

    if support:
        lines.append(f"지원 내용 : {shorten_issue_text(support)}")

    if not support and sections.get("지원 내용"):
        support_from_detail = compact_issue_values(sections["지원 내용"], max_items=1)
        if support_from_detail:
            lines.append(f"지원 내용 : {support_from_detail}")

    for label in ("구성", "요청/이슈", "원인", "조치", "기타"):
        value = compact_issue_values(sections.get(label, []))
        if value:
            lines.append(f"{label} : {value}")

    if len(lines) == 1 and sections.get("개요"):
        value = compact_issue_values(sections["개요"], max_items=2)
        if value and normalize_issue_key(value) != normalize_issue_key(support):
            lines.append(f"요약 : {value}")

    deduped = []
    seen = set()
    for line in lines:
        key = normalize_issue_key(line)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(line)
        if len(deduped) >= max_lines:
            break
    return deduped or summarize_detail(record["상세내용"], support, max_lines=max_lines)


def grouped_issue_records(issue_records: list[dict]) -> list[dict]:
    groups = []
    by_key = {}
    for record in issue_records:
        customer = str(record["고객사명"] or "").strip() or "고객사 미기재"
        support = extract_support_content(record["지원내역"])
        fallback_key = record["상세내용"] or record["지원내역"]
        key = (
            normalize_issue_key(customer),
            normalize_issue_key(support or fallback_key)[:120],
        )
        if key not in by_key:
            group = {
                "customer": customer,
                "engineers": [],
                "records": [],
            }
            by_key[key] = group
            groups.append(group)
        group = by_key[key]
        engineer = str(record["엔지니어"] or "").strip() or "담당 미기재"
        if engineer not in group["engineers"]:
            group["engineers"].append(engineer)
        group["records"].append(record)
    return groups


def format_issue_date(start_at: dt.datetime | None) -> str:
    if not start_at:
        return ""
    return f"{start_at.month}/{start_at.day}"


def issue_group_date_text(group: dict) -> str:
    dates = []
    seen = set()
    for record in group["records"]:
        start_at = record.get("작업시작일시")
        if not start_at:
            continue
        key = start_at.date()
        if key in seen:
            continue
        seen.add(key)
        dates.append(key)
    dates.sort()
    return ", ".join(f"{date.month}/{date.day}" for date in dates)


def issue_header_text(group: dict) -> str:
    engineers = ",".join(group["engineers"])
    date_text = issue_group_date_text(group)
    suffix = f" - {date_text}" if date_text else ""
    return f"- {group['customer']} ({engineers}){suffix}"


def issue_text_units(text: str) -> int:
    return sum(2 if ord(char) > 127 else 1 for char in str(text or ""))


def estimate_issue_row_height(text: str) -> float:
    if not text:
        return ISSUE_ROW_BASE_HEIGHT
    units = issue_text_units(text)
    visual_lines = max(1, (units + ISSUE_ROW_WRAP_UNITS - 1) // ISSUE_ROW_WRAP_UNITS)
    explicit_lines = str(text).count("\n") + 1
    line_count = max(visual_lines, explicit_lines)
    return min(ISSUE_ROW_MAX_HEIGHT, max(ISSUE_ROW_BASE_HEIGHT, ISSUE_ROW_BASE_HEIGHT * line_count))


def is_issue_highlight_line(text: str) -> bool:
    line = re.sub(r"^[\s.ㆍ·\-]+", "", str(text or "").strip())
    return bool(re.match(r"^(요청|요청/이슈|요청사항|조치|조치사항)\s*[:：]", line))


def truncate_text(value: str, max_chars: int = 1400) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if len(text) > max_chars:
        return text[: max_chars - 3].rstrip() + "..."
    return text


def build_llm_issue_prompt(group: dict) -> str:
    records_text = []
    for index, record in enumerate(group["records"], start=1):
        start_at = record.get("작업시작일시")
        date_text = format_month_day(start_at) if start_at else ""
        records_text.append(
            "\n".join(
                [
                    f"[지원건 {index}]",
                    f"일시: {date_text}",
                    f"고객사: {record.get('고객사명') or ''}",
                    f"엔지니어: {record.get('엔지니어') or ''}",
                    f"제품: {record.get('제품명') or ''}",
                    f"종류: {record.get('종류') or ''}",
                    f"지원내용: {truncate_text(extract_support_content(record.get('지원내역') or ''), 350)}",
                    f"상세내용: {truncate_text(record.get('상세내용') or '', 1400)}",
                ]
            )
        )
    return "\n\n".join(records_text)


def parse_llm_lines(content: str, max_lines: int = 5) -> list[str]:
    lines = []
    for raw_line in str(content or "").splitlines():
        line = raw_line.strip()
        line = re.sub(r"^\s*[-*ㆍ•]\s*", "", line)
        line = re.sub(r"^\s*\d+[\.)]\s*", "", line)
        line = re.sub(r"\s+", " ", line).strip()
        if not line:
            continue
        if len(line) > 115:
            line = line[:112].rstrip() + "..."
        lines.append(line)
        if len(lines) >= max_lines:
            break
    return lines


def call_openai_chat(config: dict, prompt: str) -> str:
    url = f"{config['base_url']}/chat/completions"
    payload = {
        "model": config["model"],
        "messages": [
            {
                "role": "system",
                "content": (
                    "너는 한국어 주간업무보고 작성 보조자다. "
                    "입력된 기술지원 상세내용을 임원 보고용으로 짧고 명확하게 정리한다. "
                    "사실을 추가로 꾸며내지 말고 입력에 있는 내용만 사용한다."
                ),
            },
            {
                "role": "user",
                "content": (
                    "아래 지원건을 주간 주요 이슈에 들어갈 내용으로 요약해줘.\n"
                    "출력 규칙:\n"
                    "- 한국어로 작성\n"
                    "- 줄마다 하나의 문장 또는 항목\n"
                    "- 최대 4줄\n"
                    "- 마크다운, 번호, 따옴표 없이 출력\n"
                    "- 가능한 라벨은 '지원 내용 :', '구성 :', '요청/이슈 :', '조치 :'만 사용\n\n"
                    f"{prompt}"
                ),
            },
        ],
        "temperature": config["temperature"],
        "max_tokens": config["max_tokens"],
    }
    request = urllib.request.Request(
        url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {config['api_key']}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=config["timeout_seconds"]) as response:
        data = json.loads(response.read().decode("utf-8"))
        log_llm_usage(data, response.headers)
    return data["choices"][0]["message"]["content"]


def log_llm_usage(data: dict, headers) -> None:
    usage = data.get("usage") or {}
    prompt_tokens = int(usage.get("prompt_tokens") or usage.get("input_tokens") or 0)
    completion_tokens = int(usage.get("completion_tokens") or usage.get("output_tokens") or 0)
    total_tokens = int(usage.get("total_tokens") or prompt_tokens + completion_tokens)

    if total_tokens:
        LLM_USAGE_TOTALS["requests"] += 1
        LLM_USAGE_TOTALS["prompt_tokens"] += prompt_tokens
        LLM_USAGE_TOTALS["completion_tokens"] += completion_tokens
        LLM_USAGE_TOTALS["total_tokens"] += total_tokens
        log(
            "LLM 토큰 사용: "
            f"입력={prompt_tokens}, 출력={completion_tokens}, 합계={total_tokens}, "
            f"누적={LLM_USAGE_TOTALS['total_tokens']}"
        )

    remaining_requests = headers.get("x-ratelimit-remaining-requests")
    remaining_tokens = headers.get("x-ratelimit-remaining-tokens")
    reset_requests = headers.get("x-ratelimit-reset-requests")
    reset_tokens = headers.get("x-ratelimit-reset-tokens")
    if remaining_requests or remaining_tokens:
        log(
            "LLM rate limit 잔여: "
            f"요청={remaining_requests or 'N/A'}, 토큰={remaining_tokens or 'N/A'}, "
            f"요청리셋={reset_requests or 'N/A'}, 토큰리셋={reset_tokens or 'N/A'}"
        )


def log_llm_usage_summary() -> None:
    if not LLM_USAGE_TOTALS["requests"]:
        return
    log(
        "LLM 토큰 사용 합계: "
        f"요청={LLM_USAGE_TOTALS['requests']}, "
        f"입력={LLM_USAGE_TOTALS['prompt_tokens']}, "
        f"출력={LLM_USAGE_TOTALS['completion_tokens']}, "
        f"합계={LLM_USAGE_TOTALS['total_tokens']}"
    )


def describe_llm_error(error: Exception) -> str:
    if isinstance(error, urllib.error.HTTPError):
        reason = f"HTTP {error.code} {error.reason}".strip()
        quota_hint = " - API 한도/토큰 부족 가능성" if error.code in (402, 429) else ""
        try:
            body = error.read().decode("utf-8", errors="replace")
            body = re.sub(r"\s+", " ", body).strip()
        except Exception:
            body = ""
        if body:
            return f"{reason}{quota_hint}: {body[:300]}"
        return f"{reason}{quota_hint}"
    return type(error).__name__


def llm_issue_detail_lines(group: dict, llm_config: dict | None) -> list[str] | None:
    if not llm_config or llm_config.get("_disabled_after_error"):
        return None
    try:
        content = call_openai_chat(llm_config, build_llm_issue_prompt(group))
        lines = parse_llm_lines(content)
        return lines or None
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, KeyError, ValueError, json.JSONDecodeError) as error:
        log(f"LLM 요약 실패. 이번 실행에서는 로컬 요약으로 전환합니다: {describe_llm_error(error)}")
        llm_config["_disabled_after_error"] = True
        return None


def classify_time(start_at: dt.datetime) -> str:
    if start_at.weekday() >= 5:
        return "휴일"
    time_value = start_at.time()
    if dt.time(6, 0) <= time_value < dt.time(18, 0):
        return "주간"
    if dt.time(18, 0) <= time_value < dt.time(22, 0):
        return "야간"
    return "심야"


def format_month_day(start_at: dt.datetime | None) -> str:
    if not start_at:
        return ""
    return f"{start_at.month}월 {start_at.day}일"


def load_records(source_path: Path, expected_start: dt.date, expected_end: dt.date) -> list[dict]:
    wb = load_workbook(source_path, data_only=True, read_only=True)
    ws = wb["주간업무보고"]
    header_row, headers = find_header_row(
        ws,
        {"부서명", "엔지니어", "제품명", "지원내역", "고객사명", "담당영업", "작업시작일시"},
    )
    records = []
    out_of_range = []

    def cell(row, header):
        return row[headers[normalize_header(header)] - 1]

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not any(value is not None for value in row):
            continue
        if str(cell(row, "부서명") or "").strip() != TEAM_NAME:
            continue

        start_at = parse_datetime(cell(row, "작업시작일시"))
        if start_at:
            if start_at.date() < expected_start or start_at.date() > expected_end:
                out_of_range.append(start_at)

        content = str(cell(row, "지원내역") or "")
        kind = extract_kind(content)
        records.append(
            {
                "부서명": cell(row, "부서명"),
                "엔지니어": cell(row, "엔지니어"),
                "제품명": cell(row, "제품명"),
                "지원내역": content,
                "고객사명": cell(row, "고객사명"),
                "담당영업": cell(row, "담당영업"),
                "작업시작일시": start_at,
                "종류": kind,
                "이슈여부": has_issue_yes(content),
                "상세내용": extract_detail_content(content),
            }
        )

    if out_of_range:
        min_date = min(out_of_range).date()
        max_date = max(out_of_range).date()
        raise ValueError(
            f"다운로드 데이터가 요청 주간 밖 날짜를 포함합니다: 요청={expected_start}~{expected_end}, 실제 일부={min_date}~{max_date}"
        )

    return records


def find_label_row(ws, section_row: int, labels: list[str]) -> int:
    label_set = set(labels)
    for row in range(section_row + 1, min(ws.max_row, section_row + 12) + 1):
        values = {str(value).strip() for value in row_values(ws, row) if value not in (None, "")}
        if label_set.issubset(values):
            return row
    raise ValueError(f"{ws.title} 시트에서 라벨 행을 찾지 못했습니다: {labels}")


def write_values_under_labels(ws, label_row: int, values_by_label: dict[str, int]) -> None:
    for col in range(1, ws.max_column + 1):
        label = str(ws.cell(label_row, col).value or "").strip()
        if label in values_by_label and not isinstance(ws.cell(label_row + 1, col), MergedCell):
            ws.cell(label_row + 1, col).value = values_by_label[label]


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy.copy(source._style)
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy.copy(source.protection)


def add_row_merges_like(ws, source_row: int, target_row: int) -> None:
    ranges_to_add = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row == source_row and merged_range.max_row == source_row:
            ranges_to_add.append(
                (
                    target_row,
                    merged_range.min_col,
                    target_row,
                    merged_range.max_col,
                )
            )
    for min_row, min_col, max_row, max_col in ranges_to_add:
        coord = (
            f"{ws.cell(min_row, min_col).coordinate}:"
            f"{ws.cell(max_row, max_col).coordinate}"
        )
        if coord not in {str(rng) for rng in ws.merged_cells.ranges}:
            ws.merge_cells(coord)


def snapshot_tail(ws, start_row: int) -> dict:
    rows = []
    for row in range(start_row, ws.max_row + 1):
        cells = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cells.append(
                {
                    "value": None if isinstance(cell, MergedCell) else cell.value,
                    "style": copy.copy(cell._style) if cell.has_style else None,
                    "font": copy.copy(cell.font),
                    "fill": copy.copy(cell.fill),
                    "border": copy.copy(cell.border),
                    "alignment": copy.copy(cell.alignment),
                    "number_format": cell.number_format,
                    "protection": copy.copy(cell.protection),
                }
            )
        rows.append(
            {
                "height": ws.row_dimensions[row].height,
                "hidden": ws.row_dimensions[row].hidden,
                "cells": cells,
            }
        )

    merges = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= start_row:
            merges.append(
                (
                    merged_range.min_row - start_row,
                    merged_range.min_col,
                    merged_range.max_row - start_row,
                    merged_range.max_col,
                )
            )

    return {"rows": rows, "merges": merges}


def restore_tail(ws, target_start_row: int, snapshot: dict) -> None:
    if not snapshot["rows"]:
        return

    target_end_row = target_start_row + len(snapshot["rows"]) - 1
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.max_row >= target_start_row and merged_range.min_row <= target_end_row:
            ws.unmerge_cells(str(merged_range))

    for row_offset, row_snapshot in enumerate(snapshot["rows"]):
        row = target_start_row + row_offset
        ws.row_dimensions[row].height = row_snapshot["height"]
        ws.row_dimensions[row].hidden = row_snapshot["hidden"]
        for col, cell_snapshot in enumerate(row_snapshot["cells"], start=1):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = cell_snapshot["value"]
            if cell_snapshot["style"] is not None:
                cell._style = copy.copy(cell_snapshot["style"])
            cell.font = copy.copy(cell_snapshot["font"])
            cell.fill = copy.copy(cell_snapshot["fill"])
            cell.border = copy.copy(cell_snapshot["border"])
            cell.alignment = copy.copy(cell_snapshot["alignment"])
            cell.number_format = cell_snapshot["number_format"]
            cell.protection = copy.copy(cell_snapshot["protection"])

    for min_row, min_col, max_row, max_col in snapshot["merges"]:
        ws.merge_cells(
            start_row=target_start_row + min_row,
            start_column=min_col,
            end_row=target_start_row + max_row,
            end_column=max_col,
        )


def ensure_full_width_merge(ws, row: int, min_col: int = 1, max_col: int = 11) -> None:
    desired = (
        f"{ws.cell(row, min_col).coordinate}:"
        f"{ws.cell(row, max_col).coordinate}"
    )
    intersecting_ranges = []
    desired_exists = False
    for merged_range in list(ws.merged_cells.ranges):
        if str(merged_range) == desired:
            desired_exists = True
            continue
        if merged_range.min_row <= row <= merged_range.max_row:
            intersects = not (
                merged_range.max_col < min_col or merged_range.min_col > max_col
            )
            if intersects:
                intersecting_ranges.append(str(merged_range))
    for merged_range in intersecting_ranges:
        ws.unmerge_cells(merged_range)
    if not desired_exists:
        ws.merge_cells(desired)


def merged_ranges_overlap(left, right) -> bool:
    return not (
        left.max_row < right.min_row
        or right.max_row < left.min_row
        or left.max_col < right.min_col
        or right.max_col < left.min_col
    )


def validate_no_overlapping_merges(ws) -> None:
    ranges = list(ws.merged_cells.ranges)
    for left_index, left in enumerate(ranges):
        for right in ranges[left_index + 1 :]:
            if merged_ranges_overlap(left, right):
                raise ValueError(f"{ws.title} 시트에 겹치는 병합 범위가 있습니다: {left} / {right}")


def border_with_sides(border, left=None, right=None, top=None, bottom=None) -> Border:
    return Border(
        left=left if left is not None else copy.copy(border.left),
        right=right if right is not None else copy.copy(border.right),
        top=top if top is not None else copy.copy(border.top),
        bottom=bottom if bottom is not None else copy.copy(border.bottom),
        diagonal=copy.copy(border.diagonal),
        diagonal_direction=border.diagonal_direction,
        vertical=copy.copy(border.vertical),
        horizontal=copy.copy(border.horizontal),
        diagonalUp=border.diagonalUp,
        diagonalDown=border.diagonalDown,
        outline=border.outline,
        start=copy.copy(border.start),
        end=copy.copy(border.end),
    )


def first_border_side(ws, column: int, side_name: str):
    for row in range(1, ws.max_row + 1):
        side = getattr(ws.cell(row, column).border, side_name)
        if side and side.style:
            return copy.copy(side)
    return Side(style="medium", color="000000")


def apply_report_outer_border(ws, min_col: int = 1, max_col: int = 11) -> None:
    left_side = first_border_side(ws, min_col, "left")
    right_side = first_border_side(ws, max_col, "right")

    for row in range(1, ws.max_row + 1):
        left_cell = ws.cell(row, min_col)
        if not isinstance(left_cell, MergedCell):
            left_cell.border = border_with_sides(left_cell.border, left=left_side)

        right_cell = ws.cell(row, max_col)
        if not isinstance(right_cell, MergedCell):
            right_cell.border = border_with_sides(right_cell.border, right=right_side)

    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_col <= max_col <= merged_range.max_col:
            anchor = ws.cell(merged_range.min_row, merged_range.min_col)
            anchor.border = border_with_sides(anchor.border, right=right_side)
            merged_range.format()


def clear_section_horizontal_borders(ws, start_row: int, end_row: int, min_col: int = 1, max_col: int = 11) -> None:
    no_side = Side(style=None)
    ranges_to_reset = []

    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.max_row < start_row or merged_range.min_row > end_row:
            continue
        if merged_range.min_col > max_col or merged_range.max_col < min_col:
            continue
        ranges_to_reset.append(
            (
                str(merged_range),
                merged_range.min_row,
                merged_range.min_col,
                merged_range.max_row,
                merged_range.max_col,
            )
        )

    for coord, min_row_range, min_col_range, max_row_range, max_col_range in ranges_to_reset:
        ws.unmerge_cells(coord)
        for row in range(min_row_range, max_row_range + 1):
            for col in range(min_col_range, max_col_range + 1):
                cell = ws.cell(row, col)
                cell.border = border_with_sides(cell.border, top=no_side, bottom=no_side)
        ws.merge_cells(coord)

    for row in range(start_row, end_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            cell.border = border_with_sides(cell.border, top=no_side, bottom=no_side)

    for merged_range in ws.merged_cells.ranges:
        if merged_range.max_row < start_row or merged_range.min_row > end_row:
            continue
        if merged_range.min_col > max_col or merged_range.max_col < min_col:
            continue
        anchor = ws.cell(merged_range.min_row, merged_range.min_col)
        anchor.border = border_with_sides(anchor.border, top=no_side, bottom=no_side)
        merged_range.format()


def unmerge_rows(ws, start_row: int, end_row: int) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.max_row >= start_row and merged_range.min_row <= end_row:
            ws.unmerge_cells(str(merged_range))


def resize_section_rows(ws, section_end_row: int, first_content_row: int, needed_rows: int) -> int:
    available_rows = max(0, section_end_row - first_content_row)
    if needed_rows == available_rows:
        return section_end_row

    if needed_rows < available_rows:
        delete_start = first_content_row + needed_rows
        delete_count = available_rows - needed_rows
        unmerge_rows(ws, delete_start, section_end_row - 1)
        ws.delete_rows(delete_start, delete_count)
        return section_end_row - delete_count

    extra_rows = needed_rows - available_rows
    insert_at = section_end_row
    template_row = first_content_row
    ws.insert_rows(insert_at, extra_rows)
    for row in range(insert_at, insert_at + extra_rows):
        copy_row_style(ws, template_row, row)
        add_row_merges_like(ws, template_row, row)
        clear_row(ws, row)
    return section_end_row + extra_rows


def write_issue_rows(
    ws,
    section2_row: int,
    section3_row: int,
    issue_records: list[dict],
    llm_config: dict | None = None,
) -> None:
    first_content_row = section2_row + 1
    output_lines: list[tuple[str, bool]] = []

    if issue_records:
        output_lines.append(("", False))
    else:
        output_lines.append(("", False))

    for group in grouped_issue_records(issue_records):
        engineers = ",".join(group["engineers"])
        output_lines.append((issue_header_text(group), True))
        detail_lines = llm_issue_detail_lines(group, llm_config)
        used_local_summary = detail_lines is None
        if detail_lines is None:
            detail_lines = []
            seen = set()
            for record in group["records"]:
                for line in issue_detail_lines(record):
                    key = normalize_issue_key(line)
                    if key in seen:
                        continue
                    seen.add(key)
                    detail_lines.append(line)
                    if len(detail_lines) >= 5:
                        break
                if len(detail_lines) >= 5:
                    break
        if used_local_summary and llm_config and llm_config.get("_disabled_after_error"):
            log(f"로컬 요약 사용: {group['customer']} ({engineers})")
            for line in detail_lines:
                log(f"  - {line}")
        for line in detail_lines:
            output_lines.append((f"     . {line}", False))
        output_lines.append(("", False))

    tail_snapshot = snapshot_tail(ws, section3_row)
    section3_row = resize_section_rows(ws, section3_row, first_content_row, len(output_lines))
    restore_tail(ws, section3_row, tail_snapshot)
    for row in range(first_content_row, section3_row):
        ensure_full_width_merge(ws, row)
        clear_row(ws, row)
        ws.cell(row, 1).font = copy.copy(ws.cell(first_content_row, 1).font)
        ws.row_dimensions[row].height = 18

    for offset, (text, is_header) in enumerate(output_lines):
        row = first_content_row + offset
        cell = ws.cell(row, 1)
        cell.value = text
        ws.row_dimensions[row].height = estimate_issue_row_height(text)
        alignment = copy.copy(ws.cell(first_content_row, 1).alignment)
        alignment.wrap_text = True
        cell.alignment = alignment
        if is_header:
            font = copy.copy(ws.cell(section2_row, 1).font)
            font.bold = True
            cell.font = font
        else:
            font = copy.copy(ws.cell(first_content_row, 1).font)
            if is_issue_highlight_line(text):
                font.color = ISSUE_HIGHLIGHT_COLOR
            cell.font = font

    clear_section_horizontal_borders(ws, first_content_row, section3_row - 1)


def ensure_presales_rows(ws, header_row: int, section5_row: int, needed_rows: int) -> int:
    available_rows = max(0, section5_row - header_row - 2)
    if needed_rows <= available_rows:
        return section5_row

    extra_rows = needed_rows - available_rows
    insert_at = section5_row - 1
    template_row = header_row + 1
    tail_snapshot = snapshot_tail(ws, insert_at)
    ws.insert_rows(insert_at, extra_rows)
    for row in range(insert_at, insert_at + extra_rows):
        copy_row_style(ws, template_row, row)
        add_row_merges_like(ws, template_row, row)
        clear_row(ws, row)
    restore_tail(ws, insert_at + extra_rows, tail_snapshot)
    return section5_row + extra_rows


def remove_merged_range_safely(ws, merged_range) -> None:
    coord = str(merged_range)
    min_row = merged_range.min_row
    min_col = merged_range.min_col
    max_row = merged_range.max_row
    max_col = merged_range.max_col
    try:
        ws.unmerge_cells(coord)
    except KeyError:
        pass

    for existing_range in list(ws.merged_cells.ranges):
        if str(existing_range) == coord:
            ws.merged_cells.ranges.remove(existing_range)

    for cleanup_row in range(min_row, max_row + 1):
        for cleanup_col in range(min_col, max_col + 1):
            key = (cleanup_row, cleanup_col)
            if isinstance(ws._cells.get(key), MergedCell):
                del ws._cells[key]


def reset_presales_data_row_merges(ws, row: int) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row <= row <= merged_range.max_row:
            remove_merged_range_safely(ws, merged_range)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)


def write_presales_rows(ws, header_row: int, section5_row: int, presales_records: list[dict]) -> None:
    section5_row = ensure_presales_rows(ws, header_row, section5_row, len(presales_records))
    for row in range(header_row + 1, section5_row):
        reset_presales_data_row_merges(ws, row)
        clear_row(ws, row)

    for index, record in enumerate(presales_records, start=header_row + 1):
        current_height = ws.row_dimensions[index].height or 0
        ws.row_dimensions[index].height = max(current_height, 24)
        start_at = record["작업시작일시"]
        ws.cell(index, 1).value = format_month_day(start_at)
        ws.cell(index, 2).value = record["고객사명"]
        ws.cell(index, 3).value = "Presales"
        ws.cell(index, 4).value = extract_support_content(record["지원내역"])
        ws.cell(index, 7).value = record["엔지니어"]
        ws.cell(index, 8).value = record["담당영업"]
        ws.cell(index, 9).value = None
        ws.cell(index, 10).value = record["제품명"]


def populate_report(wb, records: list[dict]) -> dict:
    ws = wb["DS-2팀"]
    sections = find_section_rows(ws)
    issue_records = [record for record in records if record["이슈여부"]]
    llm_config = load_llm_config()
    if llm_config:
        log(f"LLM 요약 사용: provider={llm_config['provider']}, model={llm_config['model']}")
    write_issue_rows(ws, sections[2], sections[3], issue_records, llm_config)
    sections = find_section_rows(ws)

    support_records = [record for record in records if record["종류"] in KIND_LABELS]
    kind_counts = {label: 0 for label in KIND_LABELS}
    time_counts = {label: 0 for label in TIME_LABELS}
    presales_records = []

    for record in support_records:
        kind = record["종류"]
        kind_counts[kind] += 1
        if kind == "Presales":
            presales_records.append(record)
        start_at = record["작업시작일시"]
        if start_at is None:
            raise ValueError(f"작업시작일시가 비어 있어 시간대 집계를 할 수 없습니다: {record}")
        time_counts[classify_time(start_at)] += 1

    time_total = sum(time_counts[label] for label in TIME_LABELS[:-1])
    kind_total = sum(kind_counts.values())
    if time_total != kind_total:
        raise ValueError(f"집계 총합 불일치: 시간대={time_total}, 종류={kind_total}")

    time_counts["총합계"] = time_total
    kind_counts["총합계"] = kind_total

    time_label_row = find_label_row(ws, sections[3], TIME_LABELS)
    kind_label_row = find_label_row(ws, sections[3], KIND_LABELS)
    write_values_under_labels(ws, time_label_row, time_counts)
    write_values_under_labels(ws, kind_label_row, kind_counts)

    presales_header_row = find_table_header_row(ws, sections[4], sections[5])
    if presales_header_row is None:
        raise ValueError("Presales 지원사항 헤더 행을 찾지 못했습니다.")
    write_presales_rows(ws, presales_header_row, sections[5], presales_records)
    apply_middle_vertical_alignment(ws)
    apply_report_outer_border(ws)
    sections = find_section_rows(ws)
    clear_section_horizontal_borders(ws, sections[2] + 1, sections[3] - 1)

    return {
        "record_count": len(records),
        "support_record_count": len(support_records),
        "time_counts": time_counts,
        "kind_counts": kind_counts,
        "presales_count": len(presales_records),
        "issue_count": len(issue_records),
    }


def parse_args() -> argparse.Namespace:
    env_start = parse_date_arg(os.environ.get("REPORT_START_DATE"))
    env_end = parse_date_arg(os.environ.get("REPORT_END_DATE"))
    monday, sunday = this_week_range()
    default_start = (env_start or monday).strftime("%Y-%m-%d")
    default_end = (env_end or sunday).strftime("%Y-%m-%d")

    parser = argparse.ArgumentParser(description="Rockpaper 다운로드 엑셀로 DS-2팀 최종 주간보고 생성")
    parser.add_argument("--source", type=Path, default=None, help="Rockpaper 다운로드 xlsx")
    parser.add_argument(
        "--latest-download",
        action="store_true",
        help="D:\\Downloads에서 최신 주간업무보고 xlsx를 사용",
    )
    parser.add_argument("--template", type=Path, default=None, help="최종 주간보고 템플릿 xlsx")
    parser.add_argument("--output", type=Path, default=None, help="생성할 최종 보고서 xlsx")
    parser.add_argument("--start", default=default_start, help="검증 시작일 YYYY-MM-DD")
    parser.add_argument("--end", default=default_end, help="검증 종료일 YYYY-MM-DD")
    return parser.parse_args()


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    args = parse_args()
    source_path = find_latest_download() if args.latest_download or args.source is None else args.source
    template_path = args.template or find_latest_template()
    start_date = parse_date_arg(args.start)
    end_date = parse_date_arg(args.end)
    if not source_path.exists():
        raise FileNotFoundError(f"다운로드 파일을 찾지 못했습니다: {source_path}")
    if not template_path.exists():
        raise FileNotFoundError(f"템플릿 파일을 찾지 못했습니다: {template_path}")
    if start_date is None or end_date is None:
        raise ValueError("시작일/종료일은 YYYY-MM-DD 형식이어야 합니다.")
    output_path = args.output or (
        Path.cwd()
        / "outputs"
        / f"락플레이스-DS2_주간보고_{end_date.strftime('%Y%m%d')}.xlsx"
    )
    output_path = prepare_output_path(output_path)

    log(f"다운로드 파일 DS 2T 필터링: {source_path}")
    filter_downloaded_workbook(source_path)

    records = load_records(source_path, start_date, end_date)
    log(f"DS 2T 레코드 수: {len(records)}")

    wb = build_blank_report(template_path)
    summary = populate_report(wb, records)
    for worksheet in wb.worksheets:
        validate_no_overlapping_merges(worksheet)

    output_path = save_workbook_with_fallback(wb, output_path)
    log(f"최종 보고서 생성 완료: {output_path}")
    log(f"지원 대상 레코드 수: {summary['support_record_count']}")
    log(f"시간대 집계: {summary['time_counts']}")
    log(f"종류 집계: {summary['kind_counts']}")
    log(f"Presales 건수: {summary['presales_count']}")
    log(f"주간 주요 이슈 건수: {summary['issue_count']}")
    log_llm_usage_summary()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
