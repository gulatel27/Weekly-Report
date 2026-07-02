import argparse
import copy
import datetime as dt
from pathlib import Path
import re
import sys

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


KEEP_SHEETS = ["DS-2팀"]
DOWNLOADS_DIR = Path("D:/Downloads")
DEFAULT_TEMPLATE_GLOB = "락플레이스-지원부문_주간보고_*.xlsx"


def log(message: str) -> None:
    print(f"[{dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {message}")


def this_week_range(today: dt.date | None = None) -> tuple[dt.date, dt.date]:
    today = today or dt.date.today()
    monday = today - dt.timedelta(days=today.weekday())
    sunday = monday + dt.timedelta(days=6)
    return monday, sunday


def find_latest_template() -> Path:
    candidates = sorted(
        DOWNLOADS_DIR.glob(DEFAULT_TEMPLATE_GLOB),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(
            f"템플릿 파일을 찾지 못했습니다: {DOWNLOADS_DIR / DEFAULT_TEMPLATE_GLOB}"
        )
    return candidates[0]


def clear_cell(cell) -> None:
    if isinstance(cell, MergedCell):
        return
    cell.value = None


def clear_row(ws, row_number: int) -> None:
    for col in range(1, ws.max_column + 1):
        clear_cell(ws.cell(row_number, col))


def clear_whitespace_values(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.strip() == "":
                cell.value = None


def set_sheet_zoom(ws, zoom: int = 100) -> None:
    ws.sheet_view.zoomScale = zoom
    ws.sheet_view.zoomScaleNormal = zoom
    ws.sheet_view.zoomScalePageLayoutView = zoom


def normalize_section_title_numbers(ws) -> None:
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row, 1)
        if not isinstance(cell.value, str):
            continue
        cell.value = re.sub(r"^(\s*[23])\s*,", r"\1.", cell.value, count=1)


def apply_middle_vertical_alignment(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            alignment = copy.copy(cell.alignment)
            alignment.vertical = "center"
            cell.alignment = alignment


def find_section_rows(ws) -> dict[int, int]:
    sections: dict[int, int] = {}
    patterns = {
        1: re.compile(r"^\s*1\."),
        2: re.compile(r"^\s*2[\.,]"),
        3: re.compile(r"^\s*3[\.,]"),
        4: re.compile(r"^\s*4\."),
        5: re.compile(r"^\s*5\."),
        6: re.compile(r"^\s*6\."),
    }

    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if not isinstance(value, str):
            continue
        text = value.strip()
        for section_no, pattern in patterns.items():
            if section_no not in sections and pattern.search(text):
                sections[section_no] = row
                break
    return sections


def next_section_start(sections: dict[int, int], section_no: int, fallback: int) -> int:
    later = [row for no, row in sections.items() if no > section_no]
    return min(later) if later else fallback + 1


def row_values(ws, row_number: int) -> list:
    return [ws.cell(row_number, col).value for col in range(1, ws.max_column + 1)]


def is_summary_label_row(values: list) -> bool:
    non_empty = [value for value in values if value not in (None, "")]
    if not non_empty:
        return True
    return all(isinstance(value, str) and not value.startswith("=") for value in non_empty)


def is_section1_structure_row(value) -> bool:
    if not isinstance(value, str):
        return False
    text = value.strip()
    return (
        text.startswith("(1)")
        or text.startswith("(2)")
        or text.startswith("- 내부 기술 세미나")
        or text.startswith("- 자격증")
    )


def find_table_header_row(ws, start_row: int, end_row: int) -> int | None:
    for row in range(start_row + 1, end_row):
        values = row_values(ws, row)
        labels = {str(value).strip() for value in values if value not in (None, "")}
        if "일시" in labels and ("담당" in labels or "내용" in labels or "엔드유저" in labels):
            return row
    return None


def blank_report_sheet(ws) -> None:
    sections = find_section_rows(ws)
    missing = [section_no for section_no in range(1, 7) if section_no not in sections]
    if missing:
        raise ValueError(f"{ws.title} 시트에서 섹션을 찾지 못했습니다: {missing}")

    # 1. 중점 추진 과제: 하위 구조만 남기고 내용은 비웁니다.
    start = sections[1] + 1
    end = next_section_start(sections, 1, ws.max_row)
    for row in range(start, end):
        if not is_section1_structure_row(ws.cell(row, 1).value):
            clear_row(ws, row)

    # 2. 주간 주요 이슈: 자동 생성 대상이므로 제목만 남깁니다.
    start = sections[2] + 1
    end = next_section_start(sections, 2, ws.max_row)
    for row in range(start, end):
        clear_row(ws, row)

    # 3. 주간 지원 요약: 표 라벨은 남기고 수치/수식만 비웁니다.
    start = sections[3] + 1
    end = next_section_start(sections, 3, ws.max_row)
    for row in range(start, end):
        if not is_summary_label_row(row_values(ws, row)):
            clear_row(ws, row)

    # 4, 5. 표 헤더만 남깁니다.
    for section_no in (4, 5):
        start = sections[section_no] + 1
        end = next_section_start(sections, section_no, ws.max_row)
        header_row = find_table_header_row(ws, sections[section_no], end)
        for row in range(start, end):
            if row != header_row:
                clear_row(ws, row)

    # 6. 기타 보고 사항: 제목만 남깁니다.
    start = sections[6] + 1
    for row in range(start, ws.max_row + 1):
        clear_row(ws, row)


def build_form(template_path: Path, output_path: Path) -> Path:
    wb = load_workbook(template_path)

    for sheet_name in list(wb.sheetnames):
        if sheet_name not in KEEP_SHEETS:
            del wb[sheet_name]

    for sheet_name in KEEP_SHEETS:
        wb[sheet_name]["A4"] = '= "보고일자 " & TEXT(NOW(), "yyyy-mm-dd")'
        set_sheet_zoom(wb[sheet_name])
        blank_report_sheet(wb[sheet_name])
        normalize_section_title_numbers(wb[sheet_name])
        clear_whitespace_values(wb[sheet_name])
        apply_middle_vertical_alignment(wb[sheet_name])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def parse_args() -> argparse.Namespace:
    monday, sunday = this_week_range()
    default_output = (
        Path.cwd()
        / "outputs"
        / f"락플레이스-DS2_주간보고_빈폼_{sunday.strftime('%Y%m%d')}.xlsx"
    )

    parser = argparse.ArgumentParser(description="DS-2팀 주간보고 빈 폼 생성")
    parser.add_argument("--template", type=Path, default=None, help="최종 주간보고 템플릿 xlsx")
    parser.add_argument("--output", type=Path, default=default_output, help="생성할 xlsx 경로")
    return parser.parse_args()


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    args = parse_args()
    template_path = args.template or find_latest_template()
    log(f"템플릿: {template_path}")
    log(f"출력: {args.output}")

    output_path = build_form(template_path, args.output)
    log(f"생성 완료: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
